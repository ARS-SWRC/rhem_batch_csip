###################################
##
##  Purpose: Run the RHEM model in batch mode using the RHEM CSIP web service
##           hosted by CSU. Scenarios inputs and model outputs can be saved using the RHEM_template.xlsx spreadsheet.
##           This script will allow the web service to be ran asynchronously.
## 
##  Author: Gerardo Armendariz
##  

import os
import sys
import requests
import json
import time
from time import gmtime, strftime
import itertools
from openpyxl import load_workbook
from openpyxl import Workbook

import asyncio
from aiohttp import ClientSession, TCPConnector, ClientTimeout

###### MODIFY THESE VALUES TO RUN RHEM BATCH SCRIPT
###### Note: If you are planning on doing large batch runs (greater than 2,0000) please let us know. 
######       You can email gerardo.armendariz@usda.gov  
SCENARIO_COUNT = 1                    # the number of scenarios (rows) to run
OUTPUT_DIR = "output"                 # the output directory where paramter and summary files will be saved
WORKBOOK_Name = "RHEM_template.xlsx"  # the workbook used for inputs and results
###################################################

try:
    RHEM_WORKBOOK = load_workbook(WORKBOOK_Name,data_only=True)
except:
    print("The Excel template file was not found.")

CSIP_RHEM_URL = 'http://csip.engr.colostate.edu:8083/csip-rhem/m/rhem/runrhem/1.0'

#####
#  Main function
#
def main():
    # create an output directory
    createOutputDirectory()
    
    # start an async event loop to run RHEM for all scenarios in Excel template
    loop = asyncio.get_event_loop()
    future = asyncio.ensure_future(openAndRunRHEMScenarios())
    loop.run_until_complete(future)

####
# Creates file system repository to save RHEM scenario outputs
#
def createOutputDirectory():
    try:
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
    except:
        print("The output directory for RHEM outputs to be stored could not be created!")

####
# Opens the RHEM scenarios from the workbook, validates inputs, runs CSIP web service, and saves results
#
async def openAndRunRHEMScenarios():
    ws = RHEM_WORKBOOK.active

    # object holding all asyncrhnous runs
    tasks = []

    error_message = ""
    row_index = 1

    t0 = time.time()
    print("Begin time: " + strftime("%a, %d %b %Y %H:%M:%S", gmtime()))
    # Use the aiohttp ClientSession to asyncronously create tasks for each RHEM scenario run.
    # Throttle the async calls to 10 connections at a time.
    connector = TCPConnector(limit=10)
    timeout = ClientTimeout(total=86400)
    async with ClientSession(connector=connector,timeout=timeout) as session:
        for row in ws.iter_rows(min_row=2, max_col=20, max_row=SCENARIO_COUNT + 1):
            ## Validate for empty input values
            inputs = (row[0].value, row[2].value, row[3].value, row[4].value, row[5].value, row[7].value, row[8].value, row[9].value, row[10].value,row[11].value,row[12].value,row[13].value,row[14].value, row[15].value,row[16].value, row[17].value)
            if any(s is None for s in inputs):
                error_message = "Please make sure that you have entered all required inputs to run the current scenario"
                print(error_message)
                ws.cell(row=row_index + 1, column=25).value = error_message
                row_index = row_index + 1
                RHEM_WORKBOOK.save(WORKBOOK_Name)
                continue

            ## Validate cover values
            # total canopy cover = Bunch + Forbs + Shrubs + Sod
            total_canopy = round(float(row[10].value) + float(row[11].value) + float(row[12].value) + float(row[13].value), 2)
            # total ground cover = Basal + Rock + Litter + Biological Crusts    
            total_ground = round(float(row[14].value) + float(row[15].value) + float(row[16].value) + float(row[17].value), 2)
            
            if total_canopy > 100 or total_ground > 100:
                error_message = "Skipping scenario " + str(row[0].value) + ". Total canopy cover and total ground cover cannot exceed 100%"
                print(error_message)
                ws.cell(row=row_index + 1, column=25).value = error_message
                row_index = row_index + 1
                RHEM_WORKBOOK.save(WORKBOOK_Name)
                continue
            else:
                if ws.cell(row=row_index + 1, column=22).value is None:
                    # Validation: replace periods for underscores in scenario names
                    scenario_name = str(row[0].value).replace(".","_")
                    
                    # default SAR to 0 in order for the service to run
                    SAR = row[6].value
                    if SAR is None:   
                        SAR = 0

                    # crete the input file/request to run the curren scenario
                    request_data = createInputFile(row_index, row_index,  scenario_name, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, SAR, 25, row[7].value, row[8].value, row[9].value, row[10].value, row[11].value, row[12].value, row[13].value, row[14].value, row[15].value,row[16].value, row[17].value)
                    
                    task = asyncio.ensure_future(runRHEMCSIPServiceAsync(CSIP_RHEM_URL, request_data, session, row_index, scenario_name))
                    tasks.append(task)

                row_index = row_index + 1   

         # all the RHEM scenario run response bodies in this variable
        responses = await asyncio.gather(*tasks)
        #print(responses)

        # end timer
        t1 = time.time()
        print("End time: " + strftime("%a, %d %b %Y %H:%M:%S", gmtime()))
        total = t1 - t0
        print("Execution time: " + str(total) + " seconds")

    # close the evnet loop        
    await session.close()

####
# Fetch a new RHEM run using the given payload(requestBody).  Use the row_index to identify the scneario (from the Excel spreadsheet)
# being processed.
#
async def runRHEMCSIPServiceAsync(url, requestBody, session, row_index, scenario_name):
    ws = RHEM_WORKBOOK.active
    
    async with session.post(url, json=json.loads(requestBody)) as response:
        print("Finished scenario: " + scenario_name)
        
        try:
            rhem_run_response = await response.json()
        except:
            print("Error reading JSON response for scenario: " + scenario_name)

        if "error" in rhem_run_response["metainfo"]:
            error_message = rhem_run_response["metainfo"]["error"]
            print(error_message)
            ws.cell(row=row_index + 1, column=25).value = error_message
        else:
            saveScenarioParameterFile(rhem_run_response)
            saveScenarioSummaryResults(rhem_run_response)
            saveScenarioSummaryResultsToExcel(rhem_run_response, row_index)

        RHEM_WORKBOOK.save(WORKBOOK_Name)
        return rhem_run_response

####
# Saves the input parameter file
#
def saveScenarioParameterFile(rhem_run_response):
    rhem_parameters_result_url = rhem_run_response["result"][16]["value"]
    rhem_parameters_result = requests.get(rhem_parameters_result_url)
    with open(os.path.join(OUTPUT_DIR,rhem_run_response["result"][16]["name"]), 'wb') as file:
        file.write(str(rhem_parameters_result.text).encode())


####
# Saves the response summary files
#
def saveScenarioSummaryResults(rhem_run_response):
    # summary file
    rhem_summary_result_url = rhem_run_response["result"][18]["value"]
    rhem_summary_result = requests.get(rhem_summary_result_url)
    with open(os.path.join(OUTPUT_DIR,rhem_run_response["result"][18]["name"]), 'wb') as file:
        file.write(str(rhem_summary_result.text).encode())


####
# Saves the simulation output summary results to the Excel sheet
#
def saveScenarioSummaryResultsToExcel(rhem_run_response,row_index):
    ws = RHEM_WORKBOOK.active

    # save the TDS (total dissolved solids) value
    tds = rhem_run_response["result"][15]["value"]
    ws.cell(row=row_index + 1, column=24).value = str(tds)

    with open(os.path.join(OUTPUT_DIR, rhem_run_response["result"][18]["name"]), 'rb') as file:
        index = 3
        for line in itertools.islice(file, 2, 6):
            if index == 3: # Avg. Precipitation
                ws.cell(row=row_index + 1, column=20).value = str(line).split("=")[1].rstrip("'").strip(r'\n')
            elif index == 4: # Avg. Runoff
                ws.cell(row=row_index + 1, column=21).value = str(line).split("=")[1].rstrip("'").strip(r'\n')
            elif index == 5: # Avg. Soil Loss
                ws.cell(row=row_index + 1, column=22).value = str(line).split("=")[1].rstrip("'").strip(r'\n')
            elif index == 6: # Avg. Sediment Yield
                ws.cell(row=row_index + 1, column=23).value = str(line).split("=")[1].rstrip("'").strip(r'\n')
            index = index + 1
    
    RHEM_WORKBOOK.save(WORKBOOK_Name)

#####
#  Creates the input request for the CSIP RHEM service
#
def createInputFile(AoAID, rhem_site_id, scenarioname, scenariodescription, units, stateid, climatestationid, soiltexture, sar, soilmoisture, slopelength, slopeshape, slopesteepness, bunchgrasscanopycover, forbscanopycover, shrubscanopycover, sodgrasscanopycover, basalcover, rockcover, littercover, cryptogamscover):
    request_data = '''{
        "metainfo": {},
        "parameter": [
            {
                "name": "AoAID",
                "description": "Area of Analysis Identifier",
                "value": ''' + str(AoAID)  + '''
            },
            {
                "name": "rhem_site_id",
                "description": "RHEM Evaluation Site Identifier",
                "value": ''' + str(rhem_site_id) + '''
            },
            {
                "name": "scenarioname",
                "description": "RHEM Scenario Name",
                "value": "''' + str(scenarioname) + '''"
            },
            {
                "name": "scenariodescription",
                "description": "RHEM Scenario description",
                "value": "''' + str(scenariodescription) + '''"
            },
            {
                "name": "units",
                "description": "RHEM Scenario Unit of Measure, 1 = metric and 2 = English",
                "value": ''' + str(units) + '''
            },
            {
                "name": "stateid",
                "description": " State Abbreviation",
                "value": "''' + str(stateid) + '''"
            },
            {
                "name": "climatestationid",
                "description": "Climate Station Identification Number",
                "value": "''' + str(climatestationid) + '''"
            },
            {
                "name": "soiltexture",
                "description": "Surface Soil Texture Class Label",
                "value": "''' + str(soiltexture) + '''"
            },
            {
                "name": "moisturecontent",
                "description": "",
                "value": ''' + str(soilmoisture) + '''
            },
            {
                "name": "slopelength",
                "description": "Slope Length",
                "value": ''' + str(slopelength) + ''',
                "unit": "m"
            },
            {
                "name": "slopeshape",
                "description": "Slope Shape",
                "value": "''' + str(slopeshape) + '''"
            },
            {
                "name": "slopesteepness",
                "description": "Slope Steepness",
                "value": ''' + str(slopesteepness) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "bunchgrasscanopycover",
                "description": "Bunchgrass Foliar Cover Percent",
                "value": ''' + str(bunchgrasscanopycover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "forbscanopycover",
                "description": "Forbs and Annuals Foliar Cover Percent",
                "value": ''' + str(forbscanopycover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "shrubscanopycover",
                "description": "Shrub Foliar Cover Percent",
                "value": ''' + str(shrubscanopycover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "sodgrasscanopycover",
                "description": "Sodgrass Foliar Cover Percent",
                "value": ''' + str(sodgrasscanopycover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "basalcover",
                "description": "Plant Basal Cover Percent",
                "value": ''' + str(basalcover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "rockcover",
                "description": "Rock Cover Percent",
                "value": ''' + str(rockcover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "littercover",
                "description": "Litter Cover Percent",
                "value": ''' + str(littercover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
                "name": "cryptogamscover",
                "Description": "Cryptogam Cover Percent",
                "value": ''' + str(cryptogamscover) + ''',
                "unit": "%",
                "min": 0.01,
                "max": 100
            },
            {
            "name": "sar",
            "Description": "Sodium Adsorption Ratio",
            "value": ''' + str(sar) + ''',
            "unit": "%",
            "min": 0,
            "max": 50
            }
        ]
    }'''
    return request_data

####
# The main function
#
if __name__ == "__main__":
    sys.exit(main())
